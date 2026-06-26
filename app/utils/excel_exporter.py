"""
excel_exporter.py
=================
Generates all Excel reports for HR & BD modules.
Each function returns raw bytes — caller handles HTTP response.
"""
import io
import calendar
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colour palette ─────────────────────────────────────────────────────────────
DARK_BLUE  = '1A237E'
MID_BLUE   = '3949AB'
ACCENT     = '00BCD4'
GREEN      = '27AE60'
RED        = 'E74C3C'
ORANGE     = 'E67E22'
LIGHT_BLUE = 'E8EAF6'
LIGHT_GREEN= 'E8F5E9'
GREY       = 'F5F7FC'
WHITE      = 'FFFFFF'
DARK_TEXT  = '212121'
GREY_TEXT  = '757575'

# ── Style helpers ──────────────────────────────────────────────────────────────

def _header_font(size=11, bold=True, color=WHITE):
    return Font(name='Arial', size=size, bold=bold, color=color)

def _body_font(size=10, bold=False, color=DARK_TEXT):
    return Font(name='Arial', size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def _thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _title_row(ws, title, subtitle, col_count):
    """Write a merged title + subtitle at the top."""
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=col_count)
    cell = ws.cell(1, 1, title)
    cell.font      = Font(name='Arial', size=14, bold=True, color=WHITE)
    cell.fill      = _fill(DARK_BLUE)
    cell.alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=col_count)
    cell2 = ws.cell(2, 1, subtitle)
    cell2.font      = Font(name='Arial', size=10, color=GREY_TEXT)
    cell2.fill      = _fill(LIGHT_BLUE)
    cell2.alignment = _center()
    ws.row_dimensions[2].height = 18

def _header_row(ws, row, headers, bg=DARK_BLUE):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.font      = _header_font()
        c.fill      = _fill(bg)
        c.alignment = _center()
        c.border    = _thin_border()
    ws.row_dimensions[row].height = 20

def _data_row(ws, row, values, shade=False):
    bg = GREY if shade else WHITE
    for col, val in enumerate(values, 1):
        c = ws.cell(row, col, val)
        c.font      = _body_font()
        c.fill      = _fill(bg)
        c.alignment = _left()
        c.border    = _thin_border()

def _currency(val):
    return round(float(val or 0), 2)

def _to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── 1. PAYROLL REPORT ─────────────────────────────────────────────────────────

def payroll_report(payrolls, month: int, year: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Payroll Report'
    ws.freeze_panes = 'A4'

    month_name = calendar.month_name[month]
    _title_row(ws, f'Payroll Report — {month_name} {year}',
               f'Generated for {len(payrolls)} employees', 13)

    headers = [
        'Emp Code', 'Employee Name', 'Department', 'Designation',
        'Basic (₹)', 'HRA (₹)', 'Allowances (₹)', 'Gross (₹)',
        'EPF EE (₹)', 'ESIC EE (₹)', 'Prof Tax (₹)', 'Total Ded (₹)', 'Net Pay (₹)'
    ]
    _header_row(ws, 3, headers)
    _set_col_widths(ws, [10, 22, 16, 18, 11, 10, 13, 12, 11, 11, 10, 12, 12])

    for i, p in enumerate(payrolls, 4):
        emp = p.employee
        _data_row(ws, i, [
            emp.employee_code,
            emp.full_name,
            emp.department.name  if emp.department  else '—',
            emp.designation.name if emp.designation else '—',
            _currency(p.basic_salary),
            _currency(p.hra),
            _currency(p.special_allowance + p.travel_allowance),
            _currency(p.gross_salary),
            _currency(p.epf_employee),
            _currency(p.esic_employee),
            _currency(p.professional_tax),
            _currency(p.total_deductions),
            _currency(p.net_pay),
        ], shade=(i % 2 == 0))

    # Totals row
    tr = len(payrolls) + 4
    ws.cell(tr, 1, 'TOTAL').font = _header_font(color=WHITE)
    ws.cell(tr, 1).fill = _fill(MID_BLUE)
    ws.cell(tr, 1).alignment = _center()

    for col, attr in enumerate([
        'gross_salary','epf_employee','esic_employee',
        'professional_tax','total_deductions','net_pay'
    ], 8):
        total = sum(getattr(p, attr, 0) or 0 for p in payrolls)
        c = ws.cell(tr, col, round(total, 2))
        c.font = _header_font(color=WHITE)
        c.fill = _fill(MID_BLUE)
        c.alignment = _center()
        c.border = _thin_border()

    ws.row_dimensions[tr].height = 20
    return _to_bytes(wb)


# ── 2. EPF REPORT ─────────────────────────────────────────────────────────────

def epf_report(epf_records, month: int, year: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'EPF Contribution'
    ws.freeze_panes = 'A4'

    month_name = calendar.month_name[month]
    _title_row(ws, f'EPF Contribution Report — {month_name} {year}',
               'Employee Provident Fund — As per EPF Act 1952', 8)

    headers = [
        'UAN Number', 'Employee Name', 'Emp Code',
        'Basic Wages (₹)', 'EPF Employee 12% (₹)',
        'EPS Employer 8.33% (₹)', 'EPF Employer 3.67% (₹)', 'Total EPF (₹)'
    ]
    _header_row(ws, 3, headers, bg='1565C0')
    _set_col_widths(ws, [16, 22, 10, 14, 18, 20, 20, 13])

    for i, r in enumerate(epf_records, 4):
        emp = r.employee
        _data_row(ws, i, [
            r.uan_number or '—',
            emp.full_name,
            emp.employee_code,
            _currency(r.basic_wages),
            _currency(r.epf_employee),
            _currency(r.eps_employer),
            _currency(r.epf_employer),
            _currency(r.total_epf),
        ], shade=(i % 2 == 0))

    # Totals
    tr = len(epf_records) + 4
    for col, vals in [(4, 'basic_wages'),(5,'epf_employee'),
                      (6,'eps_employer'),(7,'epf_employer'),(8,'total_epf')]:
        total = sum(getattr(r, vals, 0) or 0 for r in epf_records)
        c = ws.cell(tr, col, round(total, 2))
        c.font = _header_font(color=WHITE)
        c.fill = _fill('1565C0')
        c.alignment = _center()
        c.border = _thin_border()
    ws.cell(tr, 1, 'TOTAL').font = _header_font(color=WHITE)
    ws.cell(tr, 1).fill = _fill('1565C0')
    ws.cell(tr, 1).alignment = _center()

    return _to_bytes(wb)


# ── 3. ESIC REPORT ────────────────────────────────────────────────────────────

def esic_report(esic_records, month: int, year: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'ESIC Contribution'
    ws.freeze_panes = 'A4'

    month_name = calendar.month_name[month]
    _title_row(ws, f'ESIC Contribution Report — {month_name} {year}',
               'Employee State Insurance — As per ESI Act 1948', 7)

    headers = [
        'IP Number', 'Employee Name', 'Emp Code',
        'Gross Wages (₹)', 'ESIC Employee 0.75% (₹)',
        'ESIC Employer 3.25% (₹)', 'Total ESIC (₹)'
    ]
    _header_row(ws, 3, headers, bg='1B5E20')
    _set_col_widths(ws, [16, 22, 10, 14, 22, 22, 13])

    for i, r in enumerate(esic_records, 4):
        emp = r.employee
        _data_row(ws, i, [
            r.ip_number or '—',
            emp.full_name,
            emp.employee_code,
            _currency(r.gross_wages),
            _currency(r.esic_employee),
            _currency(r.esic_employer),
            _currency(r.total_esic),
        ], shade=(i % 2 == 0))

    tr = len(esic_records) + 4
    for col, attr in [(4,'gross_wages'),(5,'esic_employee'),
                      (6,'esic_employer'),(7,'total_esic')]:
        total = sum(getattr(r, attr, 0) or 0 for r in esic_records)
        c = ws.cell(tr, col, round(total, 2))
        c.font = _header_font(color=WHITE)
        c.fill = _fill('1B5E20')
        c.alignment = _center()
        c.border = _thin_border()
    ws.cell(tr, 1, 'TOTAL').font = _header_font(color=WHITE)
    ws.cell(tr, 1).fill = _fill('1B5E20')
    ws.cell(tr, 1).alignment = _center()

    return _to_bytes(wb)


# ── 4. ATTENDANCE REPORT ──────────────────────────────────────────────────────

def attendance_report(report_data, month: int, year: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'
    ws.freeze_panes = 'A4'

    month_name = calendar.month_name[month]
    _title_row(ws, f'Attendance Report — {month_name} {year}',
               f'{len(report_data)} employees', 8)

    headers = [
        'Emp Code', 'Employee Name', 'Department',
        'Present', 'Absent', 'Half Day', 'WFH', 'Total Days'
    ]
    _header_row(ws, 3, headers, bg='006064')
    _set_col_widths(ws, [10, 22, 18, 10, 10, 10, 10, 12])

    for i, row in enumerate(report_data, 4):
        emp = row['employee']
        _data_row(ws, i, [
            emp.employee_code,
            emp.full_name,
            emp.department.name if emp.department else '—',
            row['present'],
            row['absent'],
            row['half_day'],
            row['wfh'],
            row['total'],
        ], shade=(i % 2 == 0))

        # Colour-code present/absent cells
        ws.cell(i, 4).font = Font(name='Arial', size=10,
                                   bold=True, color=GREEN)
        ws.cell(i, 5).font = Font(name='Arial', size=10,
                                   bold=True, color=RED)

    return _to_bytes(wb)


# ── 5. LEAVE REPORT ───────────────────────────────────────────────────────────

def leave_report(leaves) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Leave Report'
    ws.freeze_panes = 'A4'

    _title_row(ws, 'Leave Report — All Requests',
               f'Total {len(leaves)} leave requests', 9)

    headers = [
        'Emp Code', 'Employee Name', 'Department',
        'Leave Type', 'From Date', 'To Date', 'Days', 'Reason', 'Status'
    ]
    _header_row(ws, 3, headers, bg=ORANGE)
    _set_col_widths(ws, [10, 22, 16, 16, 12, 12, 7, 24, 10])

    STATUS_COLORS = {
        'approved': GREEN,
        'rejected': RED,
        'pending':  ORANGE,
        'voided':   '78909C',
    }

    for i, l in enumerate(leaves, 4):
        emp = l.employee
        _data_row(ws, i, [
            emp.employee_code,
            emp.full_name,
            emp.department.name if emp.department else '—',
            l.leave_type.name,
            l.from_date.strftime('%d-%b-%Y'),
            l.to_date.strftime('%d-%b-%Y'),
            l.days,
            l.reason or '—',
            l.status.upper(),
        ], shade=(i % 2 == 0))

        color = STATUS_COLORS.get(l.status, DARK_TEXT)
        ws.cell(i, 9).font = Font(name='Arial', size=10,
                                   bold=True, color=color)

    return _to_bytes(wb)


# ── 6. EMPLOYEE MASTER ────────────────────────────────────────────────────────

def employee_master(employees) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Employee Master'
    ws.freeze_panes = 'A4'

    _title_row(ws, 'Employee Master Report',
               f'Total {len(employees)} employees', 14)

    headers = [
        'Emp Code', 'First Name', 'Last Name', 'Email', 'Phone',
        'Department', 'Designation', 'Joining Date', 'Employment Type',
        'Basic (₹)', 'Gross (₹)', 'EPF?', 'ESIC?', 'Status'
    ]
    _header_row(ws, 3, headers, bg='4A148C')
    _set_col_widths(ws, [10,13,13,26,13,16,18,13,15,11,11,7,7,10])

    for i, emp in enumerate(employees, 4):
        _data_row(ws, i, [
            emp.employee_code,
            emp.first_name,
            emp.last_name,
            emp.email,
            emp.phone or '—',
            emp.department.name  if emp.department  else '—',
            emp.designation.name if emp.designation else '—',
            emp.joining_date.strftime('%d-%b-%Y'),
            emp.employment_type.replace('_', ' ').title(),
            _currency(emp.basic_salary),
            _currency(emp.gross_salary),
            'Yes' if emp.epf_applicable  else 'No',
            'Yes' if emp.esic_applicable else 'No',
            emp.status.title(),
        ], shade=(i % 2 == 0))

        # Green = active, red = inactive/resigned
        color = GREEN if emp.status == 'active' else RED
        ws.cell(i, 14).font = Font(name='Arial', size=10,
                                    bold=True, color=color)

    return _to_bytes(wb)


# ── 7. BD PIPELINE REPORT ─────────────────────────────────────────────────────

def bd_pipeline_report(opportunities) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'BD Pipeline'
    ws.freeze_panes = 'A4'

    _title_row(ws, 'BD Pipeline Report',
               f'Total {len(opportunities)} opportunities', 8)

    headers = [
        'Title', 'Client', 'Stage', 'Deal Value (₹)',
        'Probability %', 'Expected Close', 'Assigned To', 'Description'
    ]
    _header_row(ws, 3, headers, bg='004D40')
    _set_col_widths(ws, [28, 20, 14, 14, 13, 14, 15, 30])

    STAGE_COLORS = {
        'prospect':    '1565C0',
        'proposal':    'F57F17',
        'negotiation': '6A1B9A',
        'won':         '1B5E20',
        'lost':        'B71C1C',
    }

    for i, o in enumerate(opportunities, 4):
        _data_row(ws, i, [
            o.title,
            o.client.company_name if o.client else '—',
            o.stage.title(),
            _currency(o.value),
            o.probability,
            o.expected_close_date.strftime('%d-%b-%Y') if o.expected_close_date else '—',
            o.assigned_user.username if o.assigned_user else '—',
            o.description or '—',
        ], shade=(i % 2 == 0))

        color = STAGE_COLORS.get(o.stage, DARK_TEXT)
        ws.cell(i, 3).font = Font(name='Arial', size=10,
                                   bold=True, color=color)

    # Summary by stage at the bottom
    ws.cell(len(opportunities) + 6, 1, 'STAGE SUMMARY').font = \
        Font(name='Arial', size=11, bold=True, color=DARK_BLUE)

    stages = ['prospect', 'proposal', 'negotiation', 'won', 'lost']
    for j, stage in enumerate(stages, len(opportunities) + 7):
        count = sum(1 for o in opportunities if o.stage == stage)
        value = sum(o.value or 0 for o in opportunities if o.stage == stage)
        ws.cell(j, 1, stage.title()).font = _body_font(bold=True)
        ws.cell(j, 2, f'{count} deals').font = _body_font()
        ws.cell(j, 3, f'₹{value:,.0f}').font = _body_font()

    return _to_bytes(wb)
